"""End-to-end tests for SheetForge.

Run with:  python tests/test_flows.py
Covers: fixture generation, all merge modes, both compression presets,
ZIP bundling, the full Flask HTTP API, and error handling.
Exits non-zero if anything fails.
"""
from __future__ import annotations

import io
import os
import sys
import tempfile
import zipfile
from pathlib import Path

BASE = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BASE))

# ---------------------------------------------------------------------------
# Fixture builders
# ---------------------------------------------------------------------------
def make_image(width=1600, height=1200, color=(60, 120, 200), path=None):
    """A large, realistic-looking image (gradient + noise) that has room to shrink."""
    import random
    from PIL import Image
    random.seed(42)
    im = Image.new("RGB", (width, height), color)
    px = im.load()
    for x in range(width):
        t = x / width
        for y in range(height):
            noise = random.randint(-18, 18)
            px[x, y] = (int(60 + 120 * t) + noise,
                        int(120 + 60 * (1 - t)) + noise,
                        int(200 - 80 * t) + noise)
    if path:
        im.save(path, "PNG")
        return path
    buf = io.BytesIO()
    im.save(buf, "PNG")
    return buf.getvalue()


def make_xlsx(path, sheets, image=True):
    """sheets: list of (name, rows[list[list]])"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.formatting.rule import CellIsRule
    wb = Workbook()
    wb.remove(wb.active)
    img_path = None
    for i, (name, rows) in enumerate(sheets):
        ws = wb.create_sheet(title=name)
        for r, row in enumerate(rows):
            ws.append(row)
        if rows:
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="DDEEFF")
            ws.conditional_formatting.add(
                "A1:C50", CellIsRule(operator="greaterThan", formula=["100"],
                                     fill=PatternFill("solid", fgColor="FFCCCC"))
            )
        if image and i == 0:
            img_path = path.with_suffix(".img.png")
            make_image(path=img_path)
            img = XLImage(str(img_path))
            img.width = 300
            img.height = 220
            ws.add_image(img, "E2")
    wb.save(path)
    if img_path:
        img_path.unlink(missing_ok=True)


def make_csv(path, header, rows):
    path.write_text(",".join(header) + "\n" +
                    "\n".join(",".join(str(c) for c in r) for r in rows) + "\n",
                    encoding="utf-8")


def make_xls(path, sheets):
    import xlwt
    wb = xlwt.Workbook()
    for name, rows in sheets:
        ws = wb.add_sheet(name)
        for r, row in enumerate(rows):
            for c, v in enumerate(row):
                ws.write(r, c, v)
    wb.save(str(path))


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def expect(cond, msg):
    if not cond:
        print(f"  ✗ FAIL: {msg}")
        raise AssertionError(msg)


def check_xlsx_valid(path):
    """A produced .xlsx must be a valid zip AND loadable by openpyxl."""
    with zipfile.ZipFile(path) as z:
        assert z.testzip() is None, "corrupt zip entry"
        assert "[Content_Types].xml" in z.namelist()
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    return wb


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------
def test_merge_stack_union(tmp):
    from core.merge import read_tables, merge_stacked

    f1 = tmp / "sales_q1.xlsx"
    make_xlsx(f1, [("Data", [
        ["Region", "Amount", "Date"],
        ["North", 100, "2026-01-05"],
        ["South", 200, "2026-01-06"],
    ])], image=False)

    f2 = tmp / "sales_q2.xlsx"
    make_xlsx(f2, [("Data", [
        ["Region", "Manager", "Amount"],
        ["West", "Ada", 300],
    ])], image=False)

    f3 = tmp / "extra.csv"
    make_csv(f3, ["Region", "Amount"], [["East", 400]])

    tables = (read_tables(str(f1), "sales_q1.xlsx", False) +
              read_tables(str(f2), "sales_q2.xlsx", False) +
              read_tables(str(f3), "extra.csv", False))
    expect(len(tables) == 3, "three tables read")

    wb, stats = merge_stacked(tables, add_source=True, include_all_sheets=False)
    out = tmp / "out_stack.xlsx"
    wb.save(out)
    wb2 = check_xlsx_valid(out)

    ws = wb2.active
    header = [c.value for c in ws[1]]
    expect(header[0] == "Source File", f"source column first: {header}")
    expect("Region" in header and "Amount" in header and "Manager" in header,
           f"union includes all columns: {header}")
    expect(stats["output_rows"] == 4, f"4 data rows merged, got {stats['output_rows']}")
    expect(stats["columns"] == len(header), "column count matches")
    expect(ws.max_row == 5, "header + 4 rows")

    # dedupe: add an exact duplicate of the "West" row from file 2
    f4 = tmp / "dup.csv"
    make_csv(f4, ["Region", "Manager", "Amount"], [["West", "Ada", 300]])
    tables2 = tables + read_tables(str(f4), "dup.csv", False)
    wb3, stats3 = merge_stacked(tables2, add_source=True, dedupe=True)
    o3 = tmp / "out_dedupe.xlsx"
    wb3.save(o3)
    wb4 = check_xlsx_valid(o3)
    expect(stats3["duplicates_removed"] >= 1, "duplicate row removed")
    expect(wb4.active.max_row == 5, "no dup row in output")

    # common strategy: only shared columns
    wb5, stats5 = merge_stacked(tables, strategy="common", add_source=False)
    o5 = tmp / "out_common.xlsx"
    wb5.save(o5)
    wb6 = check_xlsx_valid(o5)
    h6 = [c.value for c in wb6.active[1]]
    expect("Manager" not in h6, f"common columns only: {h6}")

    # no-header mode
    wb7, _ = merge_stacked(tables, header=False)
    o7 = tmp / "out_noheader.xlsx"
    wb7.save(o7)
    check_xlsx_valid(o7)
    print("  ✓ merge stack (union/common/dedupe/no-header)")


def test_merge_sheets_mode(tmp):
    from core.merge import read_tables, merge_as_sheets

    f1 = tmp / "reports_a.xlsx"
    make_xlsx(f1, [
        ("Q1", [["Region", "Sales"], ["North", 10]]),
        ("Q2", [["Region", "Sales"], ["North", 20]]),
    ], image=False)
    f2 = tmp / "reports_b.csv"
    make_csv(f2, ["City", "Population"], [["Lagos", 15]])

    tables = read_tables(str(f1), "reports_a.xlsx", include_all_sheets=True) + \
             read_tables(str(f2), "reports_b.csv", False)
    expect(len(tables) == 3, "3 tables incl. both sheets")

    wb, stats = merge_as_sheets(tables, include_all_sheets=True)
    out = tmp / "out_sheets.xlsx"
    wb.save(out)
    wb2 = check_xlsx_valid(out)
    expect(wb2.sheetnames == ["reports_a_Q1", "reports_a_Q2", "reports_b"],
           f"sheet names: {wb2.sheetnames}")
    expect(stats["sheets"] == 3, "3 sheets in output")
    expect(wb2["reports_a_Q1"]["A2"].value == "North", "data preserved")
    print("  ✓ merge sheets mode (per-file sheets)")


def test_merge_xls(tmp):
    from core.merge import read_tables, merge_stacked
    f = tmp / "legacy.xls"
    make_xls(f, [("Sheet1", [
        ["Name", "Value"],
        ["Alpha", 1],
        ["Beta", 2],
    ])])
    tables = read_tables(str(f), "legacy.xls", False)
    expect(len(tables) == 1 and len(tables[0].rows) == 3, "xls table read")
    wb, stats = merge_stacked(tables)
    out = tmp / "out_xls.xlsx"
    wb.save(out)
    wb2 = check_xlsx_valid(out)
    expect(wb2.active["A2"].value == "Alpha", "xls values merged")
    expect(stats["output_rows"] == 2, "2 xls data rows")
    print("  ✓ legacy .xls support")


def test_compress_safe(tmp):
    from core.compress import compress_workbook

    src = tmp / "big.xlsx"
    make_xlsx(src, [
        ("Data", [["A", "B", "C"]] + [[f"r{i}", i * 3.14, f"text {i}"] for i in range(500)]),
        ("Empty", []),
    ], image=True)
    src_size = src.stat().st_size

    dst = tmp / "big_compressed.xlsx"
    stats = compress_workbook(str(src), str(dst), preset="safe", max_dim=800, jpeg_quality=60)

    expect(stats["images"] == 1, "image detected")
    expect(stats["images_compressed"] == 1, "image recompressed")
    expect(dst.stat().st_size < src_size, f"file shrank: {src_size} -> {dst.stat().st_size}")
    expect(stats["percent_saved"] > 0, "positive savings")

    wb = check_xlsx_valid(dst)
    ws = wb["Data"]
    expect(ws.max_row == 501, "all data rows preserved")
    expect(ws["C2"].value == "text 0", "cell values preserved")
    expect(wb.sheetnames == ["Data", "Empty"], "safe preset keeps empty sheet")
    # images preserved
    with zipfile.ZipFile(dst) as z:
        media = [n for n in z.namelist() if n.startswith("xl/media/")]
    expect(len(media) >= 1, f"images still embedded: {media}")
    print(f"  ✓ compress safe ({src_size} -> {dst.stat().st_size} bytes)")


def test_compress_max(tmp):
    from core.compress import compress_workbook

    src = tmp / "max.xlsx"
    make_xlsx(src, [
        ("Data", [["X", "Y"]] + [[i, i * 2] for i in range(200)]),
        ("BlankSheet", []),
    ], image=True)
    dst = tmp / "max_compressed.xlsx"
    stats = compress_workbook(str(src), str(dst), preset="max")

    wb = check_xlsx_valid(dst)
    expect(wb.sheetnames == ["Data"], f"empty sheet removed: {wb.sheetnames}")
    expect(stats["empty_sheets_removed"] == ["BlankSheet"], "stats record removal")
    expect(wb["Data"]["B3"].value == 2, "values intact")
    expect(dst.stat().st_size <= src.stat().st_size, "never grows")

    # the sheet PART must actually be gone, and no stale rels/content-types
    with zipfile.ZipFile(dst) as z:
        names = z.namelist()
        ct = z.read("[Content_Types].xml").decode()
    expect(not any("sheet2.xml" in n or "BlankSheet" in n for n in names),
           f"sheet part removed from package: {[n for n in names if 'sheet' in n]}")
    expect("worksheets/sheet2.xml" not in ct, "no stale content-type override")
    wbrels = None
    with zipfile.ZipFile(dst) as z:
        if "xl/_rels/workbook.xml.rels" in z.namelist():
            wbrels = z.read("xl/_rels/workbook.xml.rels").decode()
    if wbrels:
        expect("sheet2" not in wbrels, "no stale relationship to removed sheet")
    print(f"  ✓ compress max (empty sheet removed, removed parts: {stats['removed_parts']})")


def test_compress_no_growth(tmp):
    from core.compress import compress_workbook
    src = tmp / "small.xlsx"
    make_xlsx(src, [("S", [["a", "b"], [1, 2]])], image=False)
    dst = tmp / "small_c.xlsx"
    stats = compress_workbook(str(src), str(dst), preset="safe")
    expect(dst.stat().st_size <= src.stat().st_size,
           "output never larger than input")
    check_xlsx_valid(dst)
    print(f"  ✓ compress never grows a file ({stats['saved_bytes']} saved)")


def test_zip_bundle(tmp):
    from core.bundle import make_bundle
    a = tmp / "one.xlsx"
    make_xlsx(a, [("S", [["a"], [1]])], image=False)
    b = tmp / "one.xlsx"  # duplicate name on purpose
    b.write_bytes(a.read_bytes())

    dst = tmp / "bundle.zip"
    stats = make_bundle([(str(a), "one.xlsx"), (str(b), "one.xlsx")], str(dst))
    with zipfile.ZipFile(dst) as z:
        names = z.namelist()
        expect(names == ["one.xlsx", "one (2).xlsx"], f"renamed duplicates: {names}")
    expect(stats["files"] == 2, "two files bundled")
    print("  ✓ zip bundle with duplicate-name handling")


def test_flask_api(tmp):
    import json as _json
    from app import app

    client = app.test_client()

    f1 = tmp / "api_a.xlsx"
    make_xlsx(f1, [("Data", [["Col", "Val"], ["x", 1]])], image=False)
    f2 = tmp / "api_b.csv"
    make_csv(f2, ["Col", "Val"], [["y", 2]])

    # -- merge end to end -----------------------------------------------
    with open(f1, "rb") as a, open(f2, "rb") as b:
        resp = client.post("/api/merge", data={
            "files": [(a, "api_a.xlsx"), (b, "api_b.csv")],
            "mode": "stack", "strategy": "union",
            "header": "1", "add_source": "1", "dedupe": "0", "include_all": "0",
        }, content_type="multipart/form-data")
    expect(resp.status_code == 200, f"merge status {resp.status_code}: {resp.data[:300]}")
    body = _json.loads(resp.data)
    expect(body["ok"] and body["stats"]["output_rows"] == 2, "merge response stats")
    dl = client.get(body["download"])
    expect(dl.status_code == 200, "download works")
    expect(dl.data[:2] == b"PK", "downloaded file is an xlsx package")
    expect(body["name"].endswith("_merged.xlsx"), "download name")

    # -- compress end to end ---------------------------------------------
    with open(f1, "rb") as a:
        resp = client.post("/api/compress", data={
            "files": [(a, "api_a.xlsx")], "preset": "safe",
            "max_dim": "1200", "quality": "65",
        }, content_type="multipart/form-data")
    expect(resp.status_code == 200, f"compress status {resp.status_code}")
    body = _json.loads(resp.data)
    expect(body["ok"] and len(body["results"]) == 1, "compress response")
    r = client.get(body["results"][0]["download"])
    expect(r.status_code == 200 and r.data[:2] == b"PK", "compressed download")

    # -- zip end to end ----------------------------------------------------
    with open(f1, "rb") as a, open(f2, "rb") as b:
        resp = client.post("/api/zip", data={
            "files": [(a, "api_a.xlsx"), (b, "api_b.csv")],
        }, content_type="multipart/form-data")
    expect(resp.status_code == 200, "zip status")
    body = _json.loads(resp.data)
    r = client.get(body["download"])
    expect(r.status_code == 200 and r.data[:2] == b"PK", "zip download")
    with zipfile.ZipFile(io.BytesIO(r.data)) as z:
        expect(z.namelist() == ["api_a.xlsx", "api_b.csv"], "zip contents")

    # -- errors -------------------------------------------------------------
    resp = client.post("/api/merge", data={"mode": "stack"},
                       content_type="multipart/form-data")
    expect(resp.status_code == 400 and b"No files" in resp.data, "missing files error")

    bad = tmp / "bad.xlsx"
    bad.write_bytes(b"this is not a zip file at all")
    with open(bad, "rb") as a:
        resp = client.post("/api/merge", data={"files": [(a, "bad.xlsx")]},
                           content_type="multipart/form-data")
    expect(resp.status_code == 400, f"corrupt file rejected: {resp.status_code}")
    expect(b"not a valid" in resp.data, "friendly corrupt-file message")

    evil = tmp / "evil.txt"
    evil.write_text("hi")
    with open(evil, "rb") as a:
        resp = client.post("/api/zip", data={"files": [(a, "evil.txt")]},
                           content_type="multipart/form-data")
    expect(resp.status_code == 400 and b"Unsupported" in resp.data, "extension rejected")

    resp = client.get("/download/deadbeef")
    expect(resp.status_code == 404, "bad token 404")

    resp = client.get("/api/health")
    expect(resp.status_code == 200, "health check")
    print("  ✓ full HTTP API flow + error handling")


# ---------------------------------------------------------------------------
def main():
    passed = 0
    with tempfile.TemporaryDirectory() as td:
        tmp = Path(td)
        tests = [
            ("merge stack", test_merge_stack_union),
            ("merge sheets", test_merge_sheets_mode),
            ("legacy xls", test_merge_xls),
            ("compress safe", test_compress_safe),
            ("compress max", test_compress_max),
            ("compress no-growth", test_compress_no_growth),
            ("zip bundle", test_zip_bundle),
            ("flask api", test_flask_api),
        ]
        for name, fn in tests:
            try:
                print(f"▶ {name}")
                fn(tmp)
                passed += 1
            except Exception as exc:
                print(f"  ✗ FAILED: {exc}")
                import traceback; traceback.print_exc()
                print(f"\n❌ {passed}/{len(tests)} tests passed")
                sys.exit(1)
    print(f"\n✅ All {passed} test groups passed — SheetForge is ready to deploy.")


if __name__ == "__main__":
    main()

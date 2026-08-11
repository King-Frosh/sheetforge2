"""Merging of Excel/CSV tables into a single workbook.

Two merge modes:

* ``stack``  — rows from every uploaded file are combined into ONE sheet,
  with columns aligned by header name (union / common / first-file).
* ``sheets`` — every uploaded table becomes its own sheet in one workbook,
  keeping each file's layout intact.

Memory design: rows are kept as plain lists and mapped to output columns
with a precomputed index map — no per-row dictionaries — so large merges
stay well inside the memory of free hosting plans (e.g. Render free's
512 MB). Very large merges additionally switch to openpyxl's streaming
(write-only) writer.
"""
from __future__ import annotations

import csv
import io
import os
import re
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from .errors import ExcelToolError

# ---------------------------------------------------------------------------
# Limits (bounded so the server cannot be OOM'd; env-overridable)
# ---------------------------------------------------------------------------
MAX_TOTAL_ROWS = int(os.environ.get("MAX_TOTAL_ROWS", "10000000"))
MAX_OUTPUT_COLS = 500         # max columns in a stacked output
MAX_SHEET_NAME_LEN = 31       # Excel hard limit
MAX_WIDTH_SAMPLE = 1000       # rows inspected per sheet for column widths
WRITE_ONLY_THRESHOLD = 25_000  # switch to memory-safe writer above this

try:  # optional legacy .xls support
    import xlrd  # type: ignore
except Exception:  # pragma: no cover
    xlrd = None


@dataclass
class Table:
    """One logical table extracted from an uploaded file."""
    source: str          # display name of the source file
    sheet: str           # display name of the source sheet
    rows: list           # raw rows (list of lists); first row may be a header
    extra: dict = field(default_factory=dict)


# ---------------------------------------------------------------------------
# Small value helpers
# ---------------------------------------------------------------------------
def _clean(value):
    """Normalise a cell: whitespace-strip strings, map empties to None."""
    if value is None:
        return None
    if isinstance(value, str):
        s = value.strip()
        return s if s else None
    return value


def _norm_key(header):
    """Header matching key: lowercase, whitespace-collapsed."""
    if header is None:
        return ""
    s = str(header).strip().lower()
    return re.sub(r"\s+", " ", s)


def _dedupe_key(value):
    """Hashable, tolerant representation of a cell for duplicate detection.

    Normalises so that 300 (number) and "300" (CSV string) match, and
    "300" vs " 300 " match.
    """
    if value is None:
        return ("none",)
    if isinstance(value, str):
        s = value.strip().lower()
        if not s:
            return ("none",)
        try:
            f = float(s)
            if f.is_integer():
                return ("num", int(f))
            return ("num", f)
        except (ValueError, OverflowError):
            return ("str", s)
    if isinstance(value, bool):
        return ("bool", value)
    if isinstance(value, (int, float)):  # numbers of any kind → one bucket
        try:
            if float(value).is_integer():
                return ("num", int(value))
            return ("num", float(value))
        except (ValueError, OverflowError):
            return ("num", float(value))
    return (type(value).__name__, value)


def _row_is_empty(row) -> bool:
    return all(_clean(v) is None for v in row)


# ---------------------------------------------------------------------------
# File readers -> list[Table]
# ---------------------------------------------------------------------------
def read_tables(path: str, source_name: str, include_all_sheets: bool = False) -> list:
    """Read an uploaded file into one or more Table objects.

    ``include_all_sheets`` only applies to .xlsx/.xlsm/.xls workbooks:
    False -> only the first non-empty sheet is used; True -> every sheet.
    """
    ext = Path(path).suffix.lower()
    if ext in (".xlsx", ".xlsm"):
        return _read_xlsx(path, source_name, include_all_sheets)
    if ext == ".xls":
        return _read_xls(path, source_name, include_all_sheets)
    if ext == ".csv":
        return [Table(source_name, "CSV", _read_csv(path))]
    raise ExcelToolError(f"Unsupported file type: {ext}")


def _read_xlsx(path: str, source_name: str, include_all_sheets: bool) -> list:
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:  # corrupted / encrypted / not a workbook
        raise ExcelToolError(
            f"Could not read “{source_name}”: it is not a valid or readable "
            f"Excel workbook ({type(exc).__name__})."
        ) from exc

    tables = []
    try:
        for ws in wb.worksheets:
            rows = []
            for row in ws.iter_rows(values_only=True):
                vals = list(row)
                if not _row_is_empty(vals):
                    rows.append(vals)
            if rows:
                tables.append(Table(source_name, ws.title or "Sheet", rows))
                if not include_all_sheets:
                    break
    finally:
        wb.close()
    return tables


def _read_xls(path: str, source_name: str, include_all_sheets: bool) -> list:
    if xlrd is None:
        raise ExcelToolError(
            f"“{source_name}” is a legacy .xls file. Install the xlrd package "
            f"on the server to support it."
        )
    try:
        book = xlrd.open_workbook(str(path))
    except Exception as exc:
        raise ExcelToolError(
            f"Could not read “{source_name}”: {exc}"
        ) from exc

    tables = []
    for sh in book.sheets():
        rows = []
        for r in range(sh.nrows):
            vals = []
            for c in range(sh.ncols):
                cell = sh.cell(r, c)
                if cell.ctype == xlrd.XL_CELL_DATE:
                    vals.append(xlrd.xldate_as_datetime(cell.value, book.datemode))
                elif cell.ctype == xlrd.XL_CELL_EMPTY:
                    vals.append(None)
                elif cell.ctype == xlrd.XL_CELL_TEXT:
                    s = cell.value.strip()
                    vals.append(s if s else None)
                else:
                    vals.append(cell.value)
            if not _row_is_empty(vals):
                rows.append(vals)
        if rows:
            tables.append(Table(source_name, sh.name or "Sheet", rows))
            if not include_all_sheets:
                break
    return tables


def _read_csv(path: str) -> list:
    """Read a CSV with encoding + delimiter detection."""
    raw = Path(path).read_bytes()
    if not raw:
        return []

    text = None
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            text = raw.decode(enc)
            break
        except (UnicodeDecodeError, LookupError):
            continue
    if text is None:
        raise ExcelToolError("Could not detect the CSV text encoding.")

    try:
        dialect = csv.Sniffer().sniff(text[:8192], delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel

    rows = []
    for row in csv.reader(io.StringIO(text), dialect):
        vals = [None if v is None else v.strip() for v in row]
        if not _row_is_empty(vals):
            rows.append(vals)
    return rows


# ---------------------------------------------------------------------------
# Header / column alignment helpers
# ---------------------------------------------------------------------------
def _extract_header(rows, header: bool):
    """Return (header_row, data_rows)."""
    if not rows:
        return [], []
    if header:
        for i, r in enumerate(rows):
            if not _row_is_empty(r):
                return list(r), rows[i + 1:]
        return [], []
    # No header row: synthesise Column 1..N names
    ncols = max(len(r) for r in rows) or 0
    return [f"Column {i + 1}" for i in range(ncols)], rows


def _build_header(header_row):
    """Map a header row to (unique keys, display names)."""
    keys, display, counts = [], [], {}
    for i, h in enumerate(header_row):
        base = _norm_key(h) or f"col{i + 1}"
        display_name = h if (h is not None and str(h).strip()) else f"Column {i + 1}"
        if base in counts:
            counts[base] += 1
            key = f"{base}#{counts[base]}"
        else:
            counts[base] = 1
            key = base
        keys.append(key)
        display.append(display_name)
    return keys, display


# ---------------------------------------------------------------------------
# Stack mode
# ---------------------------------------------------------------------------
def merge_stacked(tables, *, header: bool = True, strategy: str = "union",
                  add_source: bool = False, dedupe: bool = False,
                  include_all_sheets: bool = False) -> tuple:
    """Combine all tables into a single sheet, aligning columns by header.

    Returns (workbook, stats).
    """
    if not tables:
        raise ExcelToolError("No readable data found in the uploaded files.")

    # ---- parse all tables: cleaned rows as plain lists, header keys ------
    parsed, total_data_rows = [], 0
    for t in tables:
        header_row, data_rows = _extract_header(t.rows, header)
        keys, display = _build_header(header_row)
        rows = []
        for r in data_rows:
            clean = [_clean(v) for v in r]
            if all(v is None for v in clean):
                continue
            rows.append(clean)
        if not rows:
            continue
        total_data_rows += len(rows)
        parsed.append({
            "source": t.source, "sheet": t.sheet,
            "keys": keys, "display": display, "rows": rows,
        })

    if not parsed:
        raise ExcelToolError("No data rows found in the uploaded files.")
    if total_data_rows > MAX_TOTAL_ROWS:
        raise ExcelToolError(
            f"Too many rows to merge ({total_data_rows:,} — the limit is "
            f"{MAX_TOTAL_ROWS:,}). Please split the files into smaller batches."
        )

    # ---- choose the final column set --------------------------------------
    first = parsed[0]
    if strategy == "first":
        final_keys, final_display = list(first["keys"]), list(first["display"])
    elif strategy == "common":
        final_keys = [k for k in first["keys"]
                      if all(k in p["keys"] for p in parsed[1:])]
        dmap = dict(zip(first["keys"], first["display"]))
        final_display = [dmap[k] for k in final_keys]
    else:  # union (default)
        final_keys, final_display = [], []
        for p in parsed:
            for k, d in zip(p["keys"], p["display"]):
                if k not in final_keys:
                    final_keys.append(k)
                    final_display.append(d)
    if len(final_keys) > MAX_OUTPUT_COLS:
        raise ExcelToolError(
            f"Too many columns after merging ({len(final_keys)} — the limit "
            f"is {MAX_OUTPUT_COLS})."
        )

    # ---- precompute per-table source-column -> output-column map ----------
    # (list of row indexes; None = no matching column in this table)
    for p in parsed:
        kmap = {k: i for i, k in enumerate(p["keys"])}
        p["colmap"] = [kmap[k] if k in kmap else None for k in final_keys]

    # ---- write -------------------------------------------------------------
    # Normal writer for typical merges (freeze panes, bold header, autofilter
    # all supported); memory-safe write-only writer for very large merges.
    out_header = (["Source File"] if add_source else []) + final_display
    write_only = total_data_rows > WRITE_ONLY_THRESHOLD
    if write_only:
        wb = Workbook(write_only=True)
        ws = wb.create_sheet(title="Merged Report")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Merged Report"
    ws.append(out_header)

    widths = [0] * len(out_header)
    seen, dupes, written = set(), 0, 0
    for p in parsed:
        colmap = p["colmap"]
        label = p["source"]
        if include_all_sheets and p["sheet"] and p["sheet"] != "CSV":
            label = f"{p['source']} [{p['sheet']}]"
        for row in p["rows"]:
            vals = [row[i] if i is not None else None for i in colmap]
            if dedupe:
                key = tuple(_dedupe_key(v) for v in vals)
                if key in seen:
                    dupes += 1
                    continue
                seen.add(key)
            if add_source:
                vals.insert(0, label)
            ws.append(vals)
            written += 1
            if written <= MAX_WIDTH_SAMPLE:
                for i, v in enumerate(vals):
                    if isinstance(v, str) and len(v) > widths[i]:
                        widths[i] = min(len(v), 60)

    for i, w in enumerate(widths):
        if w:
            ws.column_dimensions[get_column_letter(i + 1)].width = min(60, max(12, w + 2))

    if not write_only:
        # cosmetic touches only possible with the normal writer
        if header:
            ws.freeze_panes = "A2"
            for cell in ws[1]:
                cell.font = Font(bold=True)
        if header and written:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(out_header))}{written + 1}"

    stats = {
        "mode": "stack",
        "files": len(parsed),
        "input_rows": total_data_rows,
        "output_rows": written,
        "duplicates_removed": dupes,
        "columns": len(out_header),
        "strategy": strategy,
        "source_column": bool(add_source),
    }
    return wb, stats


# ---------------------------------------------------------------------------
# Sheets mode
# ---------------------------------------------------------------------------
def merge_as_sheets(tables, *, include_all_sheets: bool = False) -> tuple:
    """Each table becomes its own sheet inside one workbook."""
    if not tables:
        raise ExcelToolError("No readable data found in the uploaded files.")

    # total rows first, so we can pick the memory-safe writer
    total_rows = sum(
        sum(1 for r in t.rows if not _row_is_empty(r)) for t in tables
    )
    if total_rows > MAX_TOTAL_ROWS:
        raise ExcelToolError(
            f"Too many rows to merge ({total_rows:,} — the limit is "
            f"{MAX_TOTAL_ROWS:,}). Please split the files into smaller batches."
        )

    write_only = total_rows > WRITE_ONLY_THRESHOLD
    wb = Workbook(write_only=write_only)
    if not write_only:
        wb.remove(wb.active)

    used_names = set()
    written_total = 0

    for t in tables:
        rows = [r for r in t.rows if not _row_is_empty(r)]
        if not rows:
            continue  # skip empty tables instead of creating blank sheets
        stem = Path(t.source).stem or "Sheet"
        if include_all_sheets and t.sheet and t.sheet != "CSV":
            base = f"{stem}_{t.sheet}"
        else:
            base = stem
        name = _sanitize_sheet_name(base, used_names)

        ws = wb.create_sheet(title=name)
        widths = [0] * max(len(r) for r in rows)
        for i, r in enumerate(rows):
            clean = [_clean(v) for v in r]
            ws.append(clean)
            written_total += 1
            if i < MAX_WIDTH_SAMPLE:
                for j, v in enumerate(clean):
                    if isinstance(v, str) and j < len(widths) and len(v) > widths[j]:
                        widths[j] = min(len(v), 60)
        for j, w in enumerate(widths):
            if w:
                ws.column_dimensions[get_column_letter(j + 1)].width = min(60, max(12, w + 2))
        if not write_only:
            ws.freeze_panes = "A2"

    if written_total == 0:
        raise ExcelToolError("No data rows found in the uploaded files.")

    stats = {
        "mode": "sheets",
        "files": len(tables),
        "sheets": len(wb.sheetnames),
        "output_rows": written_total,
    }
    return wb, stats


def _sanitize_sheet_name(name: str, used: set) -> str:
    s = re.sub(r"[\\/*?:\[\]]", " ", name or "").strip().strip("'")
    if not s:
        s = "Sheet"
    if s.upper() == "HISTORY":  # reserved name in Excel
        s = "_" + s
    s = s[:MAX_SHEET_NAME_LEN]
    base, n = s, 2
    while s.lower() in used:
        s = f"{base[:MAX_SHEET_NAME_LEN - 3]}_{n}"
        n += 1
    used.add(s.lower())
    return s
